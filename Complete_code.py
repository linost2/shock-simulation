# 变量初始化过程包括路径、变量名
path="C:/Users/Youlin Zhu/Desktop/" #文件所在目标路径
input_filename = "pressure.profile" #原始数据文件名
mergespace_file = "pressure_c.txt"  #合并空格后文件名
parse_file = "2.xlsx"               #分列后文件名
fullfill_file="3.xlsx"              #填充后文件名
extract_file=""
Vz_extract_file=""

# 将初始文件首行连续空格转化为一个写入一个新文件mergespace_filename
import pandas as pd
import openpyxl as op
import matplotlib.pyplot as plt
def mergespace():
    f=open (path+input_filename,"r")
    f1=open(path+mergespace_file,"w")
    data=f.readlines()
    for line in data:
        if line[0]==" ":
            f1.write(line[1:])
        else:
            f1.write(line[0:])

    f.close()
    f1.close()



#数据按空格分列后转化为数值写入excel,column为表头（第一行）名称
def parse():
    df = pd.read_csv(path+mergespace_file)
    column='# Chunk-averaged data for fix pressure_profile and group v_pressurexx'
    df= df[column].str.split(" ",expand=True)
    df.iloc[2:,:]=df.iloc[2:,:].apply(pd.to_numeric)
    df1=df.drop(labels=range(0,1),axis=0)
    df1.to_excel(path+parse_file,index=False,header=0)

#填充时间，根据步长转化为ps
def fullfill():
    wb =op.load_workbook(path+parse_file)
    ws=wb["Sheet1"]
    row=ws.max_row
    for i in range(1,row+1): 
        if ws.cell(i,1).value is None:
            if ws.cell(i,2).value==1:
                ws.cell(i,1).value=ws.cell(i-1,1).value/1000 #除数为：1ps/以fs为单位的步长
            else:
                ws.cell(i,1).value=ws.cell(i-1,1).value
            if ws.cell(i+1,2).value is  None:
                print("数据填充完毕，谢谢")
                break
                
    wb.save(path+fullfill_file)

    
 
#绘图代码:某时刻下参数（应力、温度、冲击方向速度）随切片位置分布的曲线绘制
def draw_fixtime_curve():
    
    df=pd.read_excel(path+fullfill_file,sheet_name=0, header=0,index_col=0)
    goal_xaxis="Coord1"
    goal_yaxis="c_vx[1]"
    Tsp=df.idxmin()["v_pressurexx"] #定位时刻，可手动改为 XX (ps)
    xaxis=df.loc[Tsp,goal_xaxis].values
    yaxis=df.loc[Tsp,goal_yaxis].values
    plt.plot(xaxis,yaxis)
    
def draw_freesurfaceVelocity():
    axistime=[]
    axisvelocity=[]
    df=pd.read_excel(path+fullfill_file,sheet_name=0)
    k=df1.shape[0]
    for i in range(k):
    if df1.iloc[i,0]>200:
        axistime.append(df1.iloc[i-1,0])
        axisvelocity.append(df1.iloc[i-1,8])
        plt.plot(a,axisvelocity)

mergespace()
parse()
fullfill()
draw_fixtime_curve()
draw_freesurfaceVelocity()
